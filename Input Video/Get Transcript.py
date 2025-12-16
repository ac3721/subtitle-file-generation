import cv2
import numpy as np
from PIL import Image
from io import BytesIO
from openpyxl import Workbook
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.units import pixels_to_points

def find_text(img, debug = False):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # Create mask for non-white pixels (tolerate near-white, e.g., >240)
    white_mask = cv2.threshold(gray, 220, 255, cv2.THRESH_BINARY_INV)[1]
    if debug:
        cv2.imshow('Contours', white_mask)
        cv2.waitKey(0)
        cv2.destroyAllWindows()
        print(cv2.countNonZero(white_mask))
        
    non_white = cv2.bitwise_and(img, img, mask=white_mask)
    if debug:
        cv2.imshow('Non', non_white)
        cv2.waitKey(0)
        cv2.destroyAllWindows()
        
    # Convert to HSV color space
    hsv = cv2.cvtColor(non_white, cv2.COLOR_BGR2HSV)

    # Define blue color range
    lower_blue = np.array([90, 75, 75])
    upper_blue = np.array([145, 255, 255])

    mask = cv2.inRange(hsv, lower_blue, upper_blue)
    kernel = np.ones((3, 3), np.uint8)
    mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel, iterations=3)
    mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, kernel, iterations=1)

    # Extract blue text from image
    blue_text_img = cv2.bitwise_and(non_white, non_white, mask=mask)
    if debug:
        plt.imshow(cv2.cvtColor(blue_text_img, cv2.COLOR_BGR2RGB))
        plt.axis("off")
        plt.show()
        print(cv2.countNonZero(mask))
        
    return cv2.countNonZero(mask), blue_text_img, abs(cv2.countNonZero(mask) - cv2.countNonZero(white_mask)) < 4000

def find_template(image, debug= False):
    coord = None
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    # Create mask for non-white pixels (tolerate near-white, e.g., >240)
    mask = cv2.threshold(gray, 200, 255, cv2.THRESH_BINARY)[1]
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    if contours:
        largest_contour = max(contours, key=cv2.contourArea)
            
        x, y, w, h = cv2.boundingRect(largest_contour)
        padding = 15
        x = max(0, x + padding)
        y = max(0, y + padding)
        w = min(image.shape[1] - x, w - 2 * padding)
        h = min(image.shape[0] - y, h - 2 * padding)
        coord = [y, y+h, x, x+w]
        
        if debug:
            copy = image.copy()
            cv2.drawContours(copy, [largest_contour], -1, (0, 255, 0), 3)
            # cropped = image[y:y+h, x:x+w]
            
            # cv2.imwrite('cropped.jpg', cropped)
            cv2.imshow('Contours', copy)
            cv2.waitKey(0)
            cv2.destroyAllWindows()
    return coord

def save_excel(frame, ws, row, column = 'C'):
    rgb_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    pil_image = Image.fromarray(rgb_image)

    # Get dimensions
    img_height_px, img_width_px = pil_image.size[1], pil_image.size[0]

    # Convert pixels to points (96 DPI standard)
    img_height_points = pixels_to_points(img_height_px)
    img_width_points = pixels_to_points(img_width_px)

    # Set cell row height = image height, column width to fit
    ws.row_dimensions[row].height = img_height_points
    ws.column_dimensions[column].width = img_width_points / 7.5 + 10

    # Save image to buffer
    buffer = BytesIO()
    pil_image.save(buffer, format='PNG')
    buffer.seek(0)

    img = XLImage(buffer)
    img.anchor = column + str(row)
    img.width = img_width_points
    img.height = img_height_points
    ws.add_image(img)

folder = 'Input Video'
video_name = "Short.mp4"
video_path = video_name #folder + '/' + video_name
cap = cv2.VideoCapture(video_path)

fps = cap.get(cv2.CAP_PROP_FPS)
frame_index = 0

prev_state = None
prev = None
text = []
change_timestamps = []
coord = None
status = False

while True:
    ret, frame = cap.read()
    if not ret:
        break

    if prev_state is None:
        present_pixel, _ , _= find_text(frame)

        if present_pixel > 5000:
            # print(present_pixel)
            coord = find_template(frame,debug=True)
            cropped = frame[coord[0]:coord[1], coord[2]:coord[3]]

            present_pixel, filtered, _ = find_text(cropped)
            prev_state = filtered
            prev = present_pixel
            text.append(cropped)

            timestamp_sec = frame_index / fps
            change_timestamps.append(timestamp_sec)
            # print(prev, timestamp_sec)
            status = True
    else:
        cropped = frame[coord[0]:coord[1], coord[2]:coord[3]]
        present_pixel, filtered, valid = find_text(cropped)

        if not valid and present_pixel < 1000:
            if status == True:
                text.append(None)
                timestamp_sec = frame_index / fps
                change_timestamps.append(timestamp_sec)
                # print(prev, present_pixel, timestamp_sec, 'None')
                status = False

        elif valid and abs(present_pixel - prev) > 500:
            prev_state = filtered
            prev = present_pixel
            text.append(cropped)
            
            timestamp_sec = frame_index / fps
            change_timestamps.append(timestamp_sec)
            # print(prev, present_pixel, timestamp_sec)
            status = True

    frame_index += 1

cap.release()

# print("Color-change timestamps (seconds):")
# for t in change_timestamps:
#     print(t)

# for frame in text:
#     if frame is not None:
#         plt.imshow(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))
#         plt.axis("off")
#         plt.show()
#     # else:
#     #     print("None")

wb = Workbook()
ws = wb.active

for i in range(len(text)):
    if text[i] is not None:
        save_excel(text[i], ws, i+2)
        ws[f'A{i+2}'] = change_timestamps[i]
        if i+1 < len(text):
            ws[f'B{i+2}'] = change_timestamps[i+1]

wb.save('excel.xlsx')